#Takes and XL file and sorts by date, adds the number of nights the guest are staying.
#Removes the NZD beside the price to make is a Float rather than a string so it can be recoognised by Excel
#Scans for chages made by the user and saves those to another Xl file, then once the computer has done its thing,
#The user inputs are added back to the file.
#Gerneral formating, font, borders, alingment.
#Saves to Output.xlsx


#Imports
import openpyxl
import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles.borders import BORDER_THIN


#Main function
def main():
    wb3 = openpyxl.load_workbook(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")
    ws3 = wb3.active
    ws3.delete_cols(1)
    wb3.save(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")

    #saves User inputs
    def save():
        wb3 = openpyxl.load_workbook(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")
        ws3 = wb3.active
        wb4 = openpyxl.load_workbook(r"C:\Users\harry\Desktop\Rstatements\input.xlsx")
        ws4 = wb4.active
        for i in ws3["A2:A" + str(ws3.max_row)]:
            for c in i:
                print(c.value)
                ws4.cell(row=c.row, column=1).value = c.value

        for i in ws3["G2:G" + str(ws3.max_row)]:
            for v in i:
                if v.value != None:
                    ws4.cell(row=v.row, column=2).value = v.value


        for i in ws3["H2:H" + str(ws3.max_row)]:
            for v in i:
                if v.value != None:
                    ws4.cell(row=v.row, column=3).value = v.value

        for i in ws3["I2:I" + str(ws3.max_row)]:
            for v in i:
                if v.value != None:
                    ws4.cell(row=v.row, column=4).value = v.value
        for i in ws3["J2:J" + str(ws3.max_row)]:
            for v in i:
                if v.value != None:
                    ws4.cell(row=v.row, column=5).value = v.value
        for i in ws3["K2:K" + str(ws3.max_row)]:
            for v in i:
                if v.value != None:
                    ws4.cell(row=v.row, column=6).value = v.value
        wb4.save(r"C:\Users\harry\Desktop\Rstatements\input.xlsx")

        wb3.save(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")
    save()


    #Adds moo.xlsx to formbook.xlsx and sorts by date of arrival
    #Then this is added to output.xlsx overwriting everything.
    wb1 = openpyxl.load_workbook(r'C:\Users\harry\Desktop\Rstatements\moo.xlsx')
    wb2 = openpyxl.load_workbook(r"C:\Users\harry\Desktop\Rstatements\formbook.xlsx")

    ws1 = wb1.active
    ws2 = wb2.active

    ws2max = ws2.max_row
    maxcol = str(ws1.max_row)
    print(maxcol)

    x = ws2max
    y = 0
    p = 0
    d = 1
    e = False
    yy=0
    xx=1

    for r in ws1['A1:G' + maxcol]:

        for c in r:
            y += 1


            ws2.cell(row=x, column=y).value = c.value

            if e == False:

                if p / 6 == 1:
                    e = True
                    x += 1
                    p = 1
                    y = 0
            else:
                if p / 8 == 1:
                    e = True
                    x += 1
                    p = 1
                    y = 0

            p += 1
            d += 1

            print(bool(x - 90 / 7))

    wb2.save(r"C:\Users\harry\Desktop\Rstatements\formbook2.xlsx")

    xl = pd.ExcelFile(r"C:\Users\harry\Desktop\Rstatements\formbook2.xlsx")
    df = xl.parse("Sheet1")
    df = df.sort_values("Check-in")

    writer = pd.ExcelWriter(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")
    df.to_excel(writer, sheet_name='Sheet1')
    writer.save()

    wb3 = openpyxl.load_workbook(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")
    ws3 = wb3.active
    ws3.delete_cols(1)


    #Adds the numbe of nights the guest is satying by subtractiong the dat of arrival from
    #the date of departure
    def nights():
        for i in range(1, ws3.max_row - 1):
            cell = ws3.cell(row=i+1, column=5)
            e = i

            e = str(e+1)
            # print(e)

            cell.value = "=SUM(D" + e + "-C" + e + ")"



    for i in range(ws3.max_row):
        cell = ws3.cell(row=i + 1, column=2)
        nextcell = ws3.cell(row=i + 2, column=2)
        now = cell.value
        next = nextcell.value
        if next == now:
            ws3.delete_rows(i + 2)
    guest = {}


    #Defines Border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    #Formats Collumns
    def columns(name, collum):
        ws3.cell(row=1, column=collum).value = name
        ws3.cell(row=1, column=collum).font = Font(bold=True)
        ws3.cell(row=1, column=collum).alignment = Alignment(horizontal='center')
        ws3.cell(row=1, column=collum).border = thin_border





    columns("Country", 9)
    columns("Phone", 10)
    columns("Room", 11)
    columns("Notes", 12)

    #Changes Cell dimensions
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 25
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 10
    ws3.column_dimensions["D"].width = 10
    ws3.column_dimensions["I"].width = 20
    ws3.column_dimensions["J"].width = 15
    ws3.column_dimensions["K"].width = 15
    ws3.column_dimensions["L"].width = 30



    wb3.save(r"C:\Users\harry\Desktop\Rstatements\output2.xlsx")


    wb4 = openpyxl.load_workbook(r"C:\Users\harry\Desktop\Rstatements\input.xlsx")
    ws4 = wb4.active


    #Adds user inputs back to output.xlsx
    for cell in ws3["B1:B"+str(ws3.max_row)]:
        for outputValue in cell:
            name=str(outputValue.value)


            for inputNames in ws4["A1:A"+str(ws4.max_row)]:
                for inputValues in inputNames:
                    #If output name == saved name:
                    if inputValues.value == name:
                        #If cell not empty
                        if ws4.cell(row=inputValues.row, column=2).value != None:


                            #add data
                            ws3.cell(row=outputValue.row, column=8).value = ws4.cell(row=inputValues.row, column=2).value
                            ws3.cell(row=outputValue.row, column=9).value = ws4.cell(row=inputValues.row, column=3).value
                            ws3.cell(row=outputValue.row, column=10).value = ws4.cell(row=inputValues.row, column=4).value
                            ws3.cell(row=outputValue.row, column=11).value = ws4.cell(row=inputValues.row, column=5).value
                            ws3.cell(row=outputValue.row, column=12).value = ws4.cell(row=inputValues.row, column=6).value











    nights()
    wb3.save(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")


