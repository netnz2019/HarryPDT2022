#Takes the booking.com spreadsheet and formats it to the universal format
from openpyxl import load_workbook
import pyexcel as p
import os

def main():
    #opens excel files
    p.save_book_as(file_name=r'C:\Users\harry\Desktop\Rstatements\rstatement.xls',
                   dest_file_name=r'C:\Users\harry\Desktop\Rstatements\booking_com.xlsx')

    def delete_col(ws, column_id):
        ws.delete_cols(column_id)



    def tobold(list):
        for i in list:
            print(i)
            cell = ws.cell(row=1, column=i)
            cell.font = cell.font.copy(bold=True)

    #Removes the NZD after the prices and then coverts the string to a float
    def removenzd():
        for i in range(ws.max_row - 1):
            # print(i)
            cell = ws.cell(row=i + 2, column=8)
            val = cell.value

            if val == None:
                pass
            else:
                space = val.index(" ")


                nzdremoved = val[:space]
                cell.value = float(nzdremoved)

    try:
        os.remove(r'C:\Users\harry\Desktop\Rstatements\formbook.xlsx')
    except:
        pass
    #removes unnecessary columns
    wb = load_workbook(r'C:\Users\harry\Desktop\Rstatements\booking_com.xlsx')
    ws = wb.active
    delete_col(ws, 1)
    delete_col(ws, 2)
    delete_col(ws, 5)
    delete_col(ws, 6)
    delete_col(ws, 8)
    delete_col(ws, 9)

    ws.insert_cols(1)
    #Moves Columns round
    max = ws.max_row
    print(max)
    coll = "E" + str(max)
    ws.move_range("E1:" + coll, cols=-4)

    delete_col(ws, 6)
    for i in range(6):
        delete_col(ws, 9)
    du = ws.cell(row=1, column=5)
    du.value = "Nights"

    nz = ws.cell(row=2, column=8)
    print(nz.value)
    removenzd()
    tobold({1, 2, 3, 4, 5, 6, 7, 8})
    #saves file
    wb.save(r"C:\Users\harry\Desktop\Rstatements\formbook.xlsx")


