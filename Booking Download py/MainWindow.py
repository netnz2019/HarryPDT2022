#Imports
from time import *
from threading import *

from PyQt5 import QtCore, QtGui, QtWidgets

import main
import vrbo
import os
import openpyxl
from datetime import *
import mainFormat
from waiting import wait

#Open Output.xlsx sheet
wb1 = openpyxl.load_workbook(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")
ws = wb1.active


#Number of Bookings
bookings = ws.max_row

#Gets current Date
date = date.today()
list = []
for i in range(2,5):
    if ws.cell(i, 3).value == str(date):
        list.append(i)



#GUI Main Class
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout_6 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout_6.setObjectName("gridLayout_6")
        self.gridLayout_4 = QtWidgets.QGridLayout()
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setMaximumSize(QtCore.QSize(800, 600))
        self.tabWidget.setStyleSheet("background-color: rgb(193, 193, 193);")
        self.tabWidget.setObjectName("tabWidget")
        self.tab_7 = QtWidgets.QWidget()
        self.tab_7.setObjectName("tab_7")
        self.gridLayout_8 = QtWidgets.QGridLayout(self.tab_7)
        self.gridLayout_8.setObjectName("gridLayout_8")
        self.radioButton_4 = QtWidgets.QRadioButton(self.tab_7)
        self.radioButton_4.setObjectName("radioButton_4")
        self.gridLayout_8.addWidget(self.radioButton_4, 2, 0, 1, 1)
        self.pushButton = QtWidgets.QPushButton(self.tab_7)
        self.pushButton.setObjectName("pushButton")
        self.gridLayout_8.addWidget(self.pushButton, 4, 0, 1, 1)
        self.gridLayout_7 = QtWidgets.QGridLayout()
        self.gridLayout_7.setContentsMargins(-1, -1, -1, 0)
        self.gridLayout_7.setObjectName("gridLayout_7")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout()
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.can = QtWidgets.QCheckBox(self.tab_7)
        self.can.setObjectName("can")
        self.verticalLayout_5.addWidget(self.can)
        self.cus = QtWidgets.QCheckBox(self.tab_7)
        self.cus.setObjectName("cus")
        self.verticalLayout_5.addWidget(self.cus)
        self.debug = QtWidgets.QCheckBox(self.tab_7)
        self.debug.setObjectName("debug")
        self.verticalLayout_5.addWidget(self.debug)
        self.gridLayout_7.addLayout(self.verticalLayout_5, 0, 0, 1, 1)
        self.verticalLayout_8 = QtWidgets.QVBoxLayout()
        self.verticalLayout_8.setObjectName("verticalLayout_8")
        self.bookings = QtWidgets.QLabel(self.tab_7)
        self.bookings.setObjectName("bookings")
        self.verticalLayout_8.addWidget(self.bookings)
        self.open = QtWidgets.QPushButton(self.tab_7)
        self.open.setObjectName("open")
        self.verticalLayout_8.addWidget(self.open)
        self.gridLayout_7.addLayout(self.verticalLayout_8, 0, 2, 1, 1)
        self.verticalLayout_7 = QtWidgets.QVBoxLayout()
        self.verticalLayout_7.setObjectName("verticalLayout_7")
        self.verticalLayout_11 = QtWidgets.QVBoxLayout()
        self.verticalLayout_11.setObjectName("verticalLayout_11")
        self.checkin = QtWidgets.QDateEdit(self.tab_7)
        self.checkin.setObjectName("checkin")
        self.verticalLayout_11.addWidget(self.checkin)
        self.checkout = QtWidgets.QDateEdit(self.tab_7)
        self.checkout.setObjectName("checkout")
        self.verticalLayout_11.addWidget(self.checkout)
        self.verticalLayout_7.addLayout(self.verticalLayout_11)
        self.gridLayout_7.addLayout(self.verticalLayout_7, 0, 1, 1, 1)
        self.gridLayout_8.addLayout(self.gridLayout_7, 0, 0, 1, 1)
        self.radioButton_5 = QtWidgets.QRadioButton(self.tab_7)
        self.radioButton_5.setObjectName("radioButton_5")
        self.gridLayout_8.addWidget(self.radioButton_5, 3, 0, 1, 1)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setContentsMargins(-1, 40, -1, -1)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.verticalLayout_17 = QtWidgets.QVBoxLayout()
        self.verticalLayout_17.setObjectName("verticalLayout_17")
        self.label_7 = QtWidgets.QLabel(self.tab_7)
        self.label_7.setObjectName("label_7")
        self.verticalLayout_17.addWidget(self.label_7)
        self.name1 = QtWidgets.QLabel(self.tab_7)
        self.name1.setText("")
        self.name1.setObjectName("name1")
        self.verticalLayout_17.addWidget(self.name1)
        self.name2 = QtWidgets.QLabel(self.tab_7)
        self.name2.setText("")
        self.name2.setObjectName("name2")
        self.verticalLayout_17.addWidget(self.name2)
        self.name3 = QtWidgets.QLabel(self.tab_7)
        self.name3.setObjectName("name3")
        self.verticalLayout_17.addWidget(self.name3)
        self.horizontalLayout_2.addLayout(self.verticalLayout_17)
        self.verticalLayout_13 = QtWidgets.QVBoxLayout()
        self.verticalLayout_13.setObjectName("verticalLayout_13")
        self.label_10 = QtWidgets.QLabel(self.tab_7)
        self.label_10.setObjectName("label_10")
        self.verticalLayout_13.addWidget(self.label_10)
        self.room1 = QtWidgets.QLabel(self.tab_7)
        self.room1.setText("")
        self.room1.setObjectName("room1")
        self.verticalLayout_13.addWidget(self.room1)
        self.room2 = QtWidgets.QLabel(self.tab_7)
        self.room2.setText("")
        self.room2.setObjectName("room2")
        self.verticalLayout_13.addWidget(self.room2)
        self.room3 = QtWidgets.QLabel(self.tab_7)
        self.room3.setText("")
        self.room3.setObjectName("room3")
        self.verticalLayout_13.addWidget(self.room3)
        self.horizontalLayout_2.addLayout(self.verticalLayout_13)
        self.verticalLayout_19 = QtWidgets.QVBoxLayout()
        self.verticalLayout_19.setObjectName("verticalLayout_19")
        self.label_13 = QtWidgets.QLabel(self.tab_7)
        self.label_13.setObjectName("label_13")
        self.verticalLayout_19.addWidget(self.label_13)
        self.in1 = QtWidgets.QLabel(self.tab_7)
        self.in1.setText("")
        self.in1.setObjectName("in1")
        self.verticalLayout_19.addWidget(self.in1)
        self.in2 = QtWidgets.QLabel(self.tab_7)
        self.in2.setText("")
        self.in2.setObjectName("in2")
        self.verticalLayout_19.addWidget(self.in2)
        self.in3 = QtWidgets.QLabel(self.tab_7)
        self.in3.setText("")
        self.in3.setObjectName("in3")
        self.verticalLayout_19.addWidget(self.in3)
        self.horizontalLayout_2.addLayout(self.verticalLayout_19)
        self.verticalLayout_18 = QtWidgets.QVBoxLayout()
        self.verticalLayout_18.setObjectName("verticalLayout_18")
        self.label_16 = QtWidgets.QLabel(self.tab_7)
        self.label_16.setObjectName("label_16")
        self.verticalLayout_18.addWidget(self.label_16)
        self.out1 = QtWidgets.QLabel(self.tab_7)
        self.out1.setText("")
        self.out1.setObjectName("out1")
        self.verticalLayout_18.addWidget(self.out1)
        self.out2 = QtWidgets.QLabel(self.tab_7)
        self.out2.setText("")
        self.out2.setObjectName("out2")
        self.verticalLayout_18.addWidget(self.out2)
        self.out3 = QtWidgets.QLabel(self.tab_7)
        self.out3.setText("")
        self.out3.setObjectName("out3")
        self.verticalLayout_18.addWidget(self.out3)
        self.horizontalLayout_2.addLayout(self.verticalLayout_18)
        self.verticalLayout_14 = QtWidgets.QVBoxLayout()
        self.verticalLayout_14.setObjectName("verticalLayout_14")
        self.label_23 = QtWidgets.QLabel(self.tab_7)
        self.label_23.setObjectName("label_23")
        self.verticalLayout_14.addWidget(self.label_23)
        self.night1 = QtWidgets.QLabel(self.tab_7)
        self.night1.setText("")
        self.night1.setObjectName("night1")
        self.verticalLayout_14.addWidget(self.night1)
        self.night2 = QtWidgets.QLabel(self.tab_7)
        self.night2.setText("")
        self.night2.setObjectName("night2")
        self.verticalLayout_14.addWidget(self.night2)
        self.night3 = QtWidgets.QLabel(self.tab_7)
        self.night3.setText("")
        self.night3.setObjectName("night3")
        self.verticalLayout_14.addWidget(self.night3)
        self.horizontalLayout_2.addLayout(self.verticalLayout_14)
        self.verticalLayout_10 = QtWidgets.QVBoxLayout()
        self.verticalLayout_10.setObjectName("verticalLayout_10")
        self.label_19 = QtWidgets.QLabel(self.tab_7)
        self.label_19.setObjectName("label_19")
        self.verticalLayout_10.addWidget(self.label_19)
        self.price1 = QtWidgets.QLabel(self.tab_7)
        self.price1.setText("")
        self.price1.setObjectName("price1")
        self.verticalLayout_10.addWidget(self.price1)
        self.price2 = QtWidgets.QLabel(self.tab_7)
        self.price2.setText("")
        self.price2.setObjectName("price2")
        self.verticalLayout_10.addWidget(self.price2)
        self.price3 = QtWidgets.QLabel(self.tab_7)
        self.price3.setText("")
        self.price3.setObjectName("price3")
        self.verticalLayout_10.addWidget(self.price3)
        self.horizontalLayout_2.addLayout(self.verticalLayout_10)
        self.gridLayout_8.addLayout(self.horizontalLayout_2, 1, 0, 1, 1)
        self.tabWidget.addTab(self.tab_7, "")
        self.tab_8 = QtWidgets.QWidget()
        self.tab_8.setObjectName("tab_8")
        self.comboBox = QtWidgets.QComboBox(self.tab_8)
        self.comboBox.setGeometry(QtCore.QRect(30, 40, 200, 26))
        self.comboBox.setCurrentText("")
        self.comboBox.setObjectName("comboBox")
        self.label = QtWidgets.QLabel(self.tab_8)
        self.label.setGeometry(QtCore.QRect(10, 90, 91, 16))
        self.label.setStyleSheet(
            "myFrame->setStyleSheet(\".QFrame{background-color: red; border: 1px solid black; border-radius: 10px;}\");")
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.tab_8)
        self.label_2.setGeometry(QtCore.QRect(30, 130, 71, 16))
        self.label_2.setAutoFillBackground(False)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.tab_8)
        self.label_3.setGeometry(QtCore.QRect(20, 170, 81, 16))
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.tab_8)
        self.label_4.setGeometry(QtCore.QRect(30, 210, 61, 16))
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.tab_8)
        self.label_5.setGeometry(QtCore.QRect(40, 250, 51, 16))
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.tab_8)
        self.label_6.setGeometry(QtCore.QRect(110, 90, 200, 16))
        self.label_6.setObjectName("label_6")
        self.label_8 = QtWidgets.QLabel(self.tab_8)
        self.label_8.setGeometry(QtCore.QRect(110, 130, 91, 16))
        self.label_8.setObjectName("label_8")
        self.label_9 = QtWidgets.QLabel(self.tab_8)
        self.label_9.setGeometry(QtCore.QRect(110, 170, 91, 16))
        self.label_9.setObjectName("label_9")
        self.label_11 = QtWidgets.QLabel(self.tab_8)
        self.label_11.setGeometry(QtCore.QRect(110, 210, 91, 16))
        self.label_11.setObjectName("label_11")
        self.label_12 = QtWidgets.QLabel(self.tab_8)
        self.label_12.setGeometry(QtCore.QRect(110, 250, 111, 20))
        self.label_12.setObjectName("label_12")
        self.label_14 = QtWidgets.QLabel(self.tab_8)
        self.label_14.setGeometry(QtCore.QRect(290, 40, 200, 31))
        self.label_14.setStyleSheet("font: 20pt \"MS Shell Dlg 2\";\n"
                                    "font: 87 8pt \"Segoe UI Black\";\n"
                                    "font: 14pt \"MS Shell Dlg 2\";\n"
                                    "font: 63 15pt \"Sitka Heading Semibold\";")
        self.label_14.setScaledContents(True)
        self.label_14.setObjectName("label_14")
        self.lineEdit = QtWidgets.QLineEdit(self.tab_8)
        self.lineEdit.setGeometry(QtCore.QRect(310, 90, 111, 20))
        self.lineEdit.setInputMask("")
        self.lineEdit.setText("")
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.tab_8)
        self.lineEdit_2.setGeometry(QtCore.QRect(310, 130, 113, 20))
        self.lineEdit_2.setText("")
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.lineEdit_3 = QtWidgets.QLineEdit(self.tab_8)
        self.lineEdit_3.setGeometry(QtCore.QRect(310, 170, 113, 20))
        self.lineEdit_3.setText("")
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.lineEdit_4 = QtWidgets.QLineEdit(self.tab_8)
        self.lineEdit_4.setGeometry(QtCore.QRect(310, 210, 113, 20))
        self.lineEdit_4.setText("")
        self.lineEdit_4.setObjectName("lineEdit_4")
        self.lineEdit_5 = QtWidgets.QLineEdit(self.tab_8)
        self.lineEdit_5.setGeometry(QtCore.QRect(310, 250, 113, 20))
        self.lineEdit_5.setText("")
        self.lineEdit_5.setObjectName("lineEdit_5")
        self.pushButton_2 = QtWidgets.QPushButton(self.tab_8)
        self.pushButton_2.setGeometry(QtCore.QRect(350, 290, 75, 23))
        self.pushButton_2.setObjectName("pushButton_2")
        self.tabWidget.addTab(self.tab_8, "")
        self.gridLayout_4.addWidget(self.tabWidget, 0, 0, 1, 1)
        self.gridLayout_6.addLayout(self.gridLayout_4, 0, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 576, 22))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(1)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)



        #events
        # Download Button Pressed
        self.pushButton.clicked.connect(lambda: self.thread())

        #Opens Output.xlsx
        self.open.clicked.connect(lambda: self.ope())

        #Saves User Inputs to Output.xlsx
        self.pushButton_2.clicked.connect(lambda: self.saveThread())

        self.comboBox.activated.connect(lambda: self.combothread())


        #Displays the info of the Guests arrviving on the current date
        for i in list:
            num = list.index(i) + 1
            #First person in list
            if num == 1:
                self.name1.setText(ws.cell(i, 2).value)
                self.in1.setText(ws.cell(i,3).value)
                self.out1.setText(ws.cell(i, 4).value)
                self.ni = datetime.strptime(ws.cell(i, 4).value, "%Y-%m-%d")-datetime.strptime(ws.cell(i, 3).value, "%Y-%m-%d")
                self.night1.setText(str(self.ni.days))
            #second Person in list
            if num == 2:
                self.name2.setText(ws.cell(i, 2).value)
                self.in2.setText(ws.cell(i,3).value)
                self.out2.setText(ws.cell(i, 4).value)
                self.ni = datetime.strptime(ws.cell(i, 4).value, "%Y-%m-%d")-datetime.strptime(ws.cell(i, 3).value, "%Y-%m-%d")
                self.night2.setText(str(self.ni.days))
            #Thrid person in list
            if num == 3:
                self.name3.setText(ws.cell(i, 2).value)
                self.in3.setText(ws.cell(i,3).value)
                self.out3.setText(ws.cell(i, 4).value)
                self.ni = datetime.strptime(ws.cell(i, 4).value, "%Y-%m-%d")-datetime.strptime(ws.cell(i, 3).value, "%Y-%m-%d")
                self.night3.setText(str(self.ni.days))





    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.radioButton_4.setText(_translate("MainWindow", "Booking.com"))
        self.pushButton.setText(_translate("MainWindow", "PushButton"))
        self.can.setText(_translate("MainWindow", "Cancelations"))
        self.cus.setText(_translate("MainWindow", "Custom Dates"))
        self.debug.setText(_translate("MainWindow", "Headed Debugging"))
        self.bookings.setText(_translate("MainWindow", "You Currently Have:" + str(bookings)+ " bookings"))
        self.open.setText(_translate("MainWindow", "Open Bookings.xls"))
        self.radioButton_5.setText(_translate("MainWindow", "Vrbo"))
        self.label_7.setText(_translate("MainWindow", "Name:"))


        self.label_10.setText(_translate("MainWindow", "Room:"))
        self.label_13.setText(_translate("MainWindow", "Check in:"))
        self.label_16.setText(_translate("MainWindow", "Check out:"))
        self.label_23.setText(_translate("MainWindow", "Nights:"))
        self.label_19.setText(_translate("MainWindow", "Price:"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_7), _translate("MainWindow", "Download"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_8), _translate("MainWindow", "Guest Info"))



        self.label_14.setText(_translate("MainWindow", "Add Details:"))
        self.lineEdit.setPlaceholderText(_translate("MainWindow", "Country"))
        self.lineEdit_2.setPlaceholderText(_translate("MainWindow", "Room"))
        self.lineEdit_3.setPlaceholderText(_translate("MainWindow", "Price"))
        self.lineEdit_4.setPlaceholderText(_translate("MainWindow", "Phone"))
        self.lineEdit_5.setPlaceholderText(_translate("MainWindow", "Notes"))
        self.pushButton_2.setText(_translate("MainWindow", "Save"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_8), _translate("MainWindow", "Guest Info"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_7), _translate("MainWindow", "Download"))
        self.label.setText(_translate("MainWindow", "Guest Name:"))
        self.label_2.setText(_translate("MainWindow", "Check In:"))
        self.label_2.setProperty("border-radius", _translate("MainWindow", "10px"))
        self.label_3.setText(_translate("MainWindow", "Check Out:"))
        self.label_4.setText(_translate("MainWindow", "Nights:"))
        self.label_5.setText(_translate("MainWindow", "Price:"))

        #Adds Guest names to Combobox on Guest Info Page
        for i in ws["B1:B"+ str(ws.max_row)]:
            for cell in i:
                self.comboBox.addItem(cell.value)


    #Multithreading for download to prevent crashing
    def thread(self):
        t1 = Thread(target=self.Operation)
        t1.start()

    def Operation(self):
        lastupdate = os.path.getmtime(r"C:\Users\harry\Desktop\Rstatements\moo.xlsx")
        thread4 = Thread(target=main.Main(self.debug.checkState()))
        thread3 = Thread(target=vrbo.Main(self.debug.checkState()))
        thread3.start()
        thread4.start()
        threa5 = Thread(target=mainFormat.main())
        threa5.start()

    #Multithreading for saving inputs to prevent crashing
    def saveThread(self):
        t = Thread(target=self.save())
        t.start()


    #Saves user inputs to the corosponding name selected on the combobox
    def save(self):
        #Lists Through all guest names
        for i in ws["B2:B"+str(ws.max_row)]:
            for val in i:
                #If Guest name == combobox name:
                if val.value == self.comboBox.currentText():
                    country = str(self.lineEdit.text())
                    room = str(self.lineEdit_2.text())
                    price = str(self.lineEdit_3.text())
                    phone = str(self.lineEdit_4.text())
                    notes = str(self.lineEdit_5.text())

                    #Takes user input and saves it to the Xl file in the corrosponding collum in the same row as the Guest
                    if country != "":
                        if len(country) > 56:
                            self.lineEdit.setPlaceholderText("To Long")
                            self.lineEdit.setText("")
                        elif len(country) < 3:
                            self.lineEdit.setPlaceholderText("To Short")
                            self.lineEdit.setText("")
                        else:
                            ws.cell(row=val.row, column=9).value = country

                    if room != "":
                        if len(room) > 8:
                            self.lineEdit_2.setPlaceholderText("To Long")
                            self.lineEdit_2.setText("")

                        else:
                            ws.cell(row=val.row, column=11).value = room

                    if price != "":
                        if int(price) > 10000:
                            self.lineEdit_3.setPlaceholderText("To Rich")
                            self.lineEdit_3.setText("")
                        else:
                            ws.cell(row=val.row, column=8).value = int(price)
                            self.combothread()

                    if phone != "":
                        if len(phone) > 15 or len(phone) < 4:
                            self.lineEdit_4.setPlaceholderText("Invalid Number")
                            self.lineEdit_4.setText("")
                        else:
                            ws.cell(row=val.row, column=10).value = int(phone)
                    if notes != "":
                        ws.cell(row=val.row, column=12).value = notes


                    wb1.save(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")

    #Opens Output.xlsx
    def ope(self):
        os.startfile(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")

    def combothread(self):
        for i in ws["B1:B"+ str(ws.max_row)]:
            for name in i:

                if name.value == self.comboBox.currentText():
                    self.label_6.setText(name.value)
                    self.label_8.setText(ws.cell(row=name.row, column=3).value)
                    self.label_9.setText(ws.cell(row=name.row, column=4).value)
                    nights = datetime.strptime(ws.cell(row=name.row, column=4).value, '%Y-%m-%d')-datetime.strptime(ws.cell(row=name.row, column=3).value, '%Y-%m-%d')

                    print(nights)
                    self.label_11.setText(str(nights.days))

                    price = ws.cell(row=name.row, column=8).value
                    if price != None:
                        print(price)
                        self.label_12.setText(str(price)+" NZD")
                    else:
                        self.label_12.setText("")








if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()

    sys.exit(app.exec_())


