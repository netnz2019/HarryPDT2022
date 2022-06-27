#!/usr/bin/python3

import pdb
import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string

wb1 = openpyxl.load_workbook(r'C:\Users\harry\Desktop\Rstatements\moo.xlsx')
wb2 = openpyxl.load_workbook(r"C:\Users\harry\Desktop\Rstatements\formbook.xlsx")


ws1 = wb1.active
ws2 = wb2.active

ws2max = ws2.max_row
maxcol = str(ws1.max_row)
print(maxcol)

x = ws2max
y =0
p = 0
d = 1
e=False
#Copying source range values from source to destination
for r in ws1['A1:G'+maxcol]:

    for c in r:
        y += 1

        #print(c.value) #Just to see the range selected
        ws2.cell(row = x, column = y).value = c.value

        if e == False:


            if p/6 ==1:
                e=True
                x += 1
                p=1
                y = 0
        else:
            if p/8 ==1:
                e=True
                x += 1
                p=1
                y = 0












        p+=1
        d+=1

        print(bool(x - 90 / 7))

wb2.save(r"C:\Users\harry\Desktop\Rstatements\formbook2.xlsx")

xl = pd.ExcelFile(r"C:\Users\harry\Desktop\Rstatements\formbook2.xlsx")
df = xl.parse("Sheet1")
df = df.sort_values("Check-in")

writer = pd.ExcelWriter(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")
df.to_excel(writer,sheet_name='Sheet1')
writer.save()

wb3 = openpyxl.load_workbook(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")
ws3 = wb3.active

ws3.delete_cols(1)


def nights():
    for i in range(ws3.max_row - 1):
        cell = ws3.cell(row=i + 2, column=5)
        e = i + 2
        e = str(e)
        # print(e)
        cell.value = "=SUM(D" + e + "-C" + e + ")"

nights()

for i in range(ws3.max_row):
    cell = ws3.cell(row=i+1, column=2)
    nextcell = ws3.cell(row=i+2, column=2)
    now = cell.value
    next = nextcell.value
    if next == now:
        ws3.delete_rows(i+2)

wb3.save(r"C:\Users\harry\Desktop\Rstatements\output.xlsx")